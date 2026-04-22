import sqlite3
import os
import csv
import hashlib
import logging
from datetime import datetime, timedelta

DB_NAME = "idcard.db"

# ── Logging setup ────────────────────────────────────────────
logging.basicConfig(
    filename="audit.log",
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

def get_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def initialize_db():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            roll_no     TEXT    NOT NULL UNIQUE,
            department  TEXT    NOT NULL,
            year        TEXT    NOT NULL,
            contact     TEXT,
            email       TEXT,
            photo_path  TEXT,
            issue_date  TEXT    NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS staff (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    NOT NULL,
            employee_id   TEXT    NOT NULL UNIQUE,
            department    TEXT    NOT NULL,
            designation   TEXT    NOT NULL,
            contact       TEXT,
            email         TEXT,
            photo_path    TEXT,
            issue_date    TEXT    NOT NULL
        )
    """)

    # ── Audit log table ──────────────────────────────────────
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp   TEXT    NOT NULL,
            action      TEXT    NOT NULL,
            record_type TEXT    NOT NULL,
            record_id   INTEGER,
            details     TEXT
        )
    """)

    # ── Users table (Feature 04) ────────────────────────────────
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT    NOT NULL UNIQUE,
            password_hash TEXT    NOT NULL,
            role          TEXT    NOT NULL DEFAULT 'viewer',
            created_at    TEXT    NOT NULL
        )
    """)

    # Seed default admin if empty
    if cursor.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        pw_hash = hashlib.sha256("admin123".encode()).hexdigest()
        cursor.execute(
            "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
            ("admin", pw_hash, "admin", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )

    conn.commit()
    conn.close()

# ── Audit helper ─────────────────────────────────────────────

def log_action(action, record_type, record_id=None, details=""):
    """Write an entry to the audit_logs table and the audit.log file."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn = get_connection()
        conn.execute(
            "INSERT INTO audit_logs (timestamp, action, record_type, record_id, details) VALUES (?,?,?,?,?)",
            (ts, action, record_type, record_id, details),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass  # never crash the app over logging
    logging.info("[%s] %s | %s | id=%s | %s", action, record_type, record_id, details, "")

def get_audit_logs(limit=200):
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM audit_logs ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_last_login():
    """Return the timestamp string of the most recent LOGIN action, or None."""
    conn = get_connection()
    row = conn.execute(
        "SELECT timestamp FROM audit_logs WHERE action = 'LOGIN' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    return row["timestamp"] if row else None


def resequence_ids(table_name):
    """Resequence IDs to remove any gaps after a deletion."""
    conn = get_connection()
    try:
        rows = conn.execute(f"SELECT id FROM {table_name} ORDER BY id").fetchall()
        for i, row in enumerate(rows, start=1):
            if row['id'] != i:
                conn.execute(f"UPDATE {table_name} SET id = ? WHERE id = ?", (i, row['id']))
        conn.execute(f"UPDATE sqlite_sequence SET seq = (SELECT COALESCE(MAX(id), 0) FROM {table_name}) WHERE name = '{table_name}'")
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

# ── User Authentication (Feature 04) ────────────────────────

def _hash_password(password):
    """SHA-256 hash a password string."""
    return hashlib.sha256(password.encode()).hexdigest()


def authenticate_user(username, password):
    """
    Validate credentials against the users table.
    Returns (success: bool, user_dict_or_error_message).
    user_dict: {"id", "username", "role"}
    """
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM users WHERE username = ?", (username,)
    ).fetchone()
    conn.close()

    if not row:
        return False, "Invalid username or password."

    if row["password_hash"] != _hash_password(password):
        return False, "Invalid username or password."

    return True, {"id": row["id"], "username": row["username"], "role": row["role"]}


def add_user(username, password, role="viewer"):
    """Add a new user account. Returns (success, message)."""
    if role not in ("admin", "operator", "viewer"):
        return False, f"Invalid role '{role}'. Must be admin, operator, or viewer."
    try:
        conn = get_connection()
        conn.execute(
            "INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
            (username, _hash_password(password), role, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
        conn.commit()
        conn.close()
        log_action("ADD_USER", "user", details=f"username={username}, role={role}")
        return True, f"User '{username}' ({role}) created successfully."
    except sqlite3.IntegrityError:
        return False, f"Username '{username}' already exists."
    except Exception as e:
        return False, str(e)


def get_all_users():
    """Return all user accounts (without password hashes)."""
    conn = get_connection()
    rows = conn.execute("SELECT id, username, role, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def delete_user(user_id, current_username):
    """Delete a user account. Cannot delete yourself or the last admin."""
    try:
        conn = get_connection()
        user = conn.execute("SELECT username, role FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            conn.close()
            return False, "User not found."
        if user["username"] == current_username:
            conn.close()
            return False, "You cannot delete your own account."
        if user["role"] == "admin":
            admin_count = conn.execute(
                "SELECT COUNT(*) FROM users WHERE role = 'admin'"
            ).fetchone()[0]
            if admin_count <= 1:
                conn.close()
                return False, "Cannot delete the last admin account."
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        log_action("DELETE_USER", "user", user_id,
                    f"username={user['username']}, role={user['role']}")
        return True, f"User '{user['username']}' deleted."
    except Exception as e:
        return False, str(e)


def update_user_role(user_id, new_role, current_username):
    """Change a user's role. Cannot demote the last admin."""
    if new_role not in ("admin", "operator", "viewer"):
        return False, f"Invalid role '{new_role}'."
    try:
        conn = get_connection()
        user = conn.execute("SELECT username, role FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            conn.close()
            return False, "User not found."
        if user["role"] == "admin" and new_role != "admin":
            admin_count = conn.execute(
                "SELECT COUNT(*) FROM users WHERE role = 'admin'"
            ).fetchone()[0]
            if admin_count <= 1:
                conn.close()
                return False, "Cannot demote the last admin account."
        conn.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        conn.commit()
        conn.close()
        log_action("UPDATE_USER_ROLE", "user", user_id,
                    f"username={user['username']}, {user['role']} -> {new_role}")
        return True, f"User '{user['username']}' role changed to '{new_role}'."
    except Exception as e:
        return False, str(e)


# ── STUDENT CRUD ────────────────────────────────────────────

def add_student(name, roll_no, department, year, contact, email, photo_path, issue_date):
    try:
        conn = get_connection()
        cur = conn.execute("""
            INSERT INTO students (name, roll_no, department, year, contact, email, photo_path, issue_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, roll_no, department, year, contact, email, photo_path, issue_date))
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        log_action("ADD", "student", new_id, f"name={name}, roll_no={roll_no}")
        return True, "Student added successfully!"
    except sqlite3.IntegrityError:
        return False, f"Roll number '{roll_no}' already exists."
    except Exception as e:
        return False, str(e)

def get_all_students():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM students ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def search_students(query):
    conn = get_connection()
    rows = conn.execute("""
        SELECT * FROM students
        WHERE name LIKE ? OR roll_no LIKE ? OR department LIKE ?
        ORDER BY name
    """, (f"%{query}%", f"%{query}%", f"%{query}%")).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def search_students_filtered(query="", department="All", sort_col="name", sort_dir="ASC"):
    """Search students with optional department filter and sort column."""
    valid_cols = {"name", "roll_no", "department", "year", "contact", "issue_date"}
    if sort_col not in valid_cols:
        sort_col = "name"
    sort_dir = "DESC" if sort_dir.upper() == "DESC" else "ASC"

    conn = get_connection()
    params = [f"%{query}%", f"%{query}%", f"%{query}%"]
    dept_clause = ""
    if department and department != "All":
        dept_clause = " AND department = ?"
        params.append(department)

    sql = f"""
        SELECT * FROM students
        WHERE (name LIKE ? OR roll_no LIKE ? OR department LIKE ?)
        {dept_clause}
        ORDER BY {sort_col} {sort_dir}
    """
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_student_departments():
    """Return distinct department values for students."""
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT department FROM students ORDER BY department").fetchall()
    conn.close()
    return ["All"] + [r["department"] for r in rows]

def get_student_by_id(student_id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def update_student(student_id, name, roll_no, department, year, contact, email, photo_path, issue_date):
    try:
        conn = get_connection()
        conn.execute("""
            UPDATE students
            SET name=?, roll_no=?, department=?, year=?, contact=?, email=?, photo_path=?, issue_date=?
            WHERE id=?
        """, (name, roll_no, department, year, contact, email, photo_path, issue_date, student_id))
        conn.commit()
        conn.close()
        log_action("UPDATE", "student", student_id, f"name={name}, roll_no={roll_no}")
        return True, "Student updated successfully!"
    except sqlite3.IntegrityError:
        return False, f"Roll number '{roll_no}' already exists."
    except Exception as e:
        return False, str(e)

def delete_student(student_id):
    try:
        rec = get_student_by_id(student_id)
        conn = get_connection()
        conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
        conn.commit()
        conn.close()
        details = f"name={rec['name']}" if rec else ""
        log_action("DELETE", "student", student_id, details)
        resequence_ids("students")
        return True, "Student deleted successfully!"
    except Exception as e:
        return False, str(e)

# ── STAFF CRUD ───────────────────────────────────────────────

def add_staff(name, employee_id, department, designation, contact, email, photo_path, issue_date):
    try:
        conn = get_connection()
        cur = conn.execute("""
            INSERT INTO staff (name, employee_id, department, designation, contact, email, photo_path, issue_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, employee_id, department, designation, contact, email, photo_path, issue_date))
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        log_action("ADD", "staff", new_id, f"name={name}, employee_id={employee_id}")
        return True, "Staff member added successfully!"
    except sqlite3.IntegrityError:
        return False, f"Employee ID '{employee_id}' already exists."
    except Exception as e:
        return False, str(e)

def get_all_staff():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM staff ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def search_staff(query):
    conn = get_connection()
    rows = conn.execute("""
        SELECT * FROM staff
        WHERE name LIKE ? OR employee_id LIKE ? OR department LIKE ?
        ORDER BY name
    """, (f"%{query}%", f"%{query}%", f"%{query}%")).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def search_staff_filtered(query="", department="All", sort_col="name", sort_dir="ASC"):
    """Search staff with optional department filter and sort column."""
    valid_cols = {"name", "employee_id", "department", "designation", "contact", "issue_date"}
    if sort_col not in valid_cols:
        sort_col = "name"
    sort_dir = "DESC" if sort_dir.upper() == "DESC" else "ASC"

    conn = get_connection()
    params = [f"%{query}%", f"%{query}%", f"%{query}%"]
    dept_clause = ""
    if department and department != "All":
        dept_clause = " AND department = ?"
        params.append(department)

    sql = f"""
        SELECT * FROM staff
        WHERE (name LIKE ? OR employee_id LIKE ? OR department LIKE ?)
        {dept_clause}
        ORDER BY {sort_col} {sort_dir}
    """
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_staff_departments():
    """Return distinct department values for staff."""
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT department FROM staff ORDER BY department").fetchall()
    conn.close()
    return ["All"] + [r["department"] for r in rows]

def get_staff_by_id(staff_id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM staff WHERE id = ?", (staff_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def update_staff(staff_id, name, employee_id, department, designation, contact, email, photo_path, issue_date):
    try:
        conn = get_connection()
        conn.execute("""
            UPDATE staff
            SET name=?, employee_id=?, department=?, designation=?, contact=?, email=?, photo_path=?, issue_date=?
            WHERE id=?
        """, (name, employee_id, department, designation, contact, email, photo_path, issue_date, staff_id))
        conn.commit()
        conn.close()
        log_action("UPDATE", "staff", staff_id, f"name={name}, employee_id={employee_id}")
        return True, "Staff updated successfully!"
    except sqlite3.IntegrityError:
        return False, f"Employee ID '{employee_id}' already exists."
    except Exception as e:
        return False, str(e)

def delete_staff(staff_id):
    try:
        rec = get_staff_by_id(staff_id)
        conn = get_connection()
        conn.execute("DELETE FROM staff WHERE id = ?", (staff_id,))
        conn.commit()
        conn.close()
        details = f"name={rec['name']}" if rec else ""
        log_action("DELETE", "staff", staff_id, details)
        resequence_ids("staff")
        return True, "Staff member deleted successfully!"
    except Exception as e:
        return False, str(e)

# ── Expiring / Expired Cards (Feature 06) ────────────────────

def get_expiring_soon(days=60):
    """
    Return students + staff whose cards are expired OR expiring
    within `days` from today.  Each dict includes extra keys:
      record_type  ('student' | 'staff')
      identifier   (roll_no or employee_id)
    """
    from datetime import date
    cutoff = date.today() - timedelta(days=365 - days)  # issue before this → expiring
    cutoff_str = cutoff.strftime("%Y-%m-%d")

    conn = get_connection()

    students = conn.execute("""
        SELECT *, 'student' AS record_type, roll_no AS identifier
        FROM students
        WHERE issue_date <= ?
        ORDER BY issue_date ASC
    """, (cutoff_str,)).fetchall()

    staff = conn.execute("""
        SELECT *, 'staff' AS record_type, employee_id AS identifier
        FROM staff
        WHERE issue_date <= ?
        ORDER BY issue_date ASC
    """, (cutoff_str,)).fetchall()

    conn.close()
    return [dict(r) for r in students] + [dict(r) for r in staff]

# ── CSV Export / Import ──────────────────────────────────────

STUDENT_FIELDS = ["id", "name", "roll_no", "department", "year", "contact", "email", "photo_path", "issue_date"]
STAFF_FIELDS   = ["id", "name", "employee_id", "department", "designation", "contact", "email", "photo_path", "issue_date"]

def export_students_csv(filepath):
    """Export all students to a CSV file. Returns (success, message)."""
    try:
        data = get_all_students()
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=STUDENT_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        log_action("EXPORT_CSV", "student", details=f"file={filepath}, rows={len(data)}")
        return True, f"Exported {len(data)} student record(s) to:\n{filepath}"
    except Exception as e:
        return False, str(e)

def import_students_csv(filepath):
    """Import students from a CSV file. Skips duplicates. Returns (success, message)."""
    try:
        added = skipped = 0
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ok, _ = add_student(
                    row.get("name", "").strip(),
                    row.get("roll_no", "").strip(),
                    row.get("department", "").strip(),
                    row.get("year", "").strip(),
                    row.get("contact", "").strip(),
                    row.get("email", "").strip(),
                    row.get("photo_path", "").strip(),
                    row.get("issue_date", datetime.today().strftime("%Y-%m-%d")),
                )
                if ok:
                    added += 1
                else:
                    skipped += 1
        log_action("IMPORT_CSV", "student", details=f"file={filepath}, added={added}, skipped={skipped}")
        return True, f"Import complete: {added} added, {skipped} skipped (duplicates)."
    except Exception as e:
        return False, str(e)

def export_staff_csv(filepath):
    """Export all staff to a CSV file. Returns (success, message)."""
    try:
        data = get_all_staff()
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=STAFF_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        log_action("EXPORT_CSV", "staff", details=f"file={filepath}, rows={len(data)}")
        return True, f"Exported {len(data)} staff record(s) to:\n{filepath}"
    except Exception as e:
        return False, str(e)

def import_staff_csv(filepath):
    """Import staff from a CSV file. Skips duplicates. Returns (success, message)."""
    try:
        added = skipped = 0
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ok, _ = add_staff(
                    row.get("name", "").strip(),
                    row.get("employee_id", "").strip(),
                    row.get("department", "").strip(),
                    row.get("designation", "").strip(),
                    row.get("contact", "").strip(),
                    row.get("email", "").strip(),
                    row.get("photo_path", "").strip(),
                    row.get("issue_date", datetime.today().strftime("%Y-%m-%d")),
                )
                if ok:
                    added += 1
                else:
                    skipped += 1
        log_action("IMPORT_CSV", "staff", details=f"file={filepath}, added={added}, skipped={skipped}")
        return True, f"Import complete: {added} added, {skipped} skipped (duplicates)."
    except Exception as e:
        return False, str(e)

# ── Expiry / notification queries ────────────────────────────

def get_expiring_soon(days=30):
    """
    Return students and staff whose ID was issued exactly 1 year ago (±days window).
    issue_date is stored as YYYY-MM-DD.  We flag records where
    (today - issue_date) is between (365-days) and (365+days) days.
    """
    today = datetime.today().date()
    low  = today - timedelta(days=365 + days)   # issued before this → already expired
    high = today - timedelta(days=365 - days)   # issued after this  → not yet near expiry

    conn = get_connection()
    students = conn.execute(
        "SELECT name, roll_no AS identifier, department, issue_date, 'student' AS record_type"
        " FROM students WHERE date(issue_date) BETWEEN ? AND ?",
        (str(low), str(high)),
    ).fetchall()
    staff = conn.execute(
        "SELECT name, employee_id AS identifier, department, issue_date, 'staff' AS record_type"
        " FROM staff WHERE date(issue_date) BETWEEN ? AND ?",
        (str(low), str(high)),
    ).fetchall()
    conn.close()
    return [dict(r) for r in students] + [dict(r) for r in staff]

# ── Analytics queries ────────────────────────────────────────

def get_analytics_summary():
    """Return top-level counts: total students, total staff, departments."""
    conn = get_connection()
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_staff    = conn.execute("SELECT COUNT(*) FROM staff").fetchone()[0]
    dept_count     = conn.execute(
        "SELECT COUNT(DISTINCT department) FROM ("
        "  SELECT department FROM students UNION ALL SELECT department FROM staff)"
    ).fetchone()[0]
    conn.close()
    return {
        "total_students": total_students,
        "total_staff":    total_staff,
        "departments":    dept_count,
        "total":          total_students + total_staff,
    }

def get_students_per_department():
    """Return list of (department, count) tuples for students."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT department, COUNT(*) as cnt FROM students GROUP BY department ORDER BY cnt DESC"
    ).fetchall()
    conn.close()
    return [(r["department"], r["cnt"]) for r in rows]

def get_staff_per_department():
    """Return list of (department, count) tuples for staff."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT department, COUNT(*) as cnt FROM staff GROUP BY department ORDER BY cnt DESC"
    ).fetchall()
    conn.close()
    return [(r["department"], r["cnt"]) for r in rows]

def get_monthly_issuance(months=12):
    """
    Return list of (month_label, student_count, staff_count) for the last N months.
    month_label is 'YYYY-MM'.
    """
    conn = get_connection()
    from datetime import date
    today = date.today()
    results = []
    for i in range(months - 1, -1, -1):
        # compute year/month stepping back i months
        month = today.month - i
        year  = today.year
        while month <= 0:
            month += 12
            year  -= 1
        label = f"{year}-{month:02d}"
        sc = conn.execute(
            "SELECT COUNT(*) FROM students WHERE strftime('%Y-%m', issue_date) = ?", (label,)
        ).fetchone()[0]
        fc = conn.execute(
            "SELECT COUNT(*) FROM staff WHERE strftime('%Y-%m', issue_date) = ?", (label,)
        ).fetchone()[0]
        results.append((label, sc, fc))
    conn.close()
    return results

# ── Database Backup ──────────────────────────────────────────

def backup_database(backup_dir="backups"):
    """
    Copy idcard.db to <backup_dir>/idcard_YYYYMMDD_HHMMSS.db.
    Returns (success, message, filepath).
    """
    import shutil
    try:
        os.makedirs(backup_dir, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(backup_dir, f"idcard_{ts}.db")
        # Use SQLite's backup API for a safe hot-copy
        src_conn  = get_connection()
        dest_conn = sqlite3.connect(dest)
        src_conn.backup(dest_conn)
        dest_conn.close()
        src_conn.close()
        log_action("BACKUP", "database", details=f"file={dest}")
        return True, f"Backup saved to:\n{dest}", dest
    except Exception as e:
        return False, str(e), ""

# ── Excel (.xlsx) Export / Import (Feature 08) ────────────────

def export_students_xlsx(filepath):
    """Export all students to an Excel file with styled headers. Returns (success, message)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        data = get_all_students()
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Header styling
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["ID", "Name", "Roll No", "Department", "Year", "Contact", "Email", "Photo Path", "Issue Date"]
        field_keys = ["id", "name", "roll_no", "department", "year", "contact", "email", "photo_path", "issue_date"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, rec in enumerate(data, 2):
            for col_idx, key in enumerate(field_keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(key, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-size columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(filepath)
        log_action("EXPORT_XLSX", "student", details=f"file={filepath}, rows={len(data)}")
        return True, f"Exported {len(data)} student record(s) to:\n{filepath}"
    except ImportError:
        return False, "openpyxl is not installed. Run: pip install openpyxl"
    except Exception as e:
        return False, str(e)


def export_staff_xlsx(filepath):
    """Export all staff to an Excel file with styled headers. Returns (success, message)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        data = get_all_staff()
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff"

        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["ID", "Name", "Employee ID", "Department", "Designation", "Contact", "Email", "Photo Path", "Issue Date"]
        field_keys = ["id", "name", "employee_id", "department", "designation", "contact", "email", "photo_path", "issue_date"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, rec in enumerate(data, 2):
            for col_idx, key in enumerate(field_keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(key, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

        wb.save(filepath)
        log_action("EXPORT_XLSX", "staff", details=f"file={filepath}, rows={len(data)}")
        return True, f"Exported {len(data)} staff record(s) to:\n{filepath}"
    except ImportError:
        return False, "openpyxl is not installed. Run: pip install openpyxl"
    except Exception as e:
        return False, str(e)


def import_students_xlsx(filepath):
    """Import students from an Excel file. Skips duplicates. Returns (success, message)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return False, "Excel file is empty."

        # Detect headers (case-insensitive)
        raw_headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0] if h is not None]

        added = skipped = 0
        for row in rows[1:]:
            values = list(row) + [""] * (len(raw_headers) - len(row))
            rec = dict(zip(raw_headers, values))
            ok, _ = add_student(
                str(rec.get("name", "")).strip(),
                str(rec.get("roll_no", "")).strip(),
                str(rec.get("department", "")).strip(),
                str(rec.get("year", "")).strip(),
                str(rec.get("contact", "")).strip(),
                str(rec.get("email", "")).strip(),
                str(rec.get("photo_path", "")).strip(),
                str(rec.get("issue_date", datetime.today().strftime("%Y-%m-%d"))).strip(),
            )
            if ok:
                added += 1
            else:
                skipped += 1

        wb.close()
        log_action("IMPORT_XLSX", "student", details=f"file={filepath}, added={added}, skipped={skipped}")
        return True, f"Import complete: {added} added, {skipped} skipped (duplicates)."
    except ImportError:
        return False, "openpyxl is not installed. Run: pip install openpyxl"
    except Exception as e:
        return False, str(e)


def import_staff_xlsx(filepath):
    """Import staff from an Excel file. Skips duplicates. Returns (success, message)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return False, "Excel file is empty."

        raw_headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0] if h is not None]

        added = skipped = 0
        for row in rows[1:]:
            values = list(row) + [""] * (len(raw_headers) - len(row))
            rec = dict(zip(raw_headers, values))
            ok, _ = add_staff(
                str(rec.get("name", "")).strip(),
                str(rec.get("employee_id", "")).strip(),
                str(rec.get("department", "")).strip(),
                str(rec.get("designation", "")).strip(),
                str(rec.get("contact", "")).strip(),
                str(rec.get("email", "")).strip(),
                str(rec.get("photo_path", "")).strip(),
                str(rec.get("issue_date", datetime.today().strftime("%Y-%m-%d"))).strip(),
            )
            if ok:
                added += 1
            else:
                skipped += 1

        wb.close()
        log_action("IMPORT_XLSX", "staff", details=f"file={filepath}, added={added}, skipped={skipped}")
        return True, f"Import complete: {added} added, {skipped} skipped (duplicates)."
    except ImportError:
        return False, "openpyxl is not installed. Run: pip install openpyxl"
    except Exception as e:
        return False, str(e)


# ── Demo Data Generator (Feature 10) ─────────────────────────

def generate_demo_data(n_students=30, n_staff=10):
    """
    Bulk-insert realistic fake student and staff records using Faker.
    Returns (success, message) with counts of inserted records.
    """
    try:
        from faker import Faker
        import random

        fake = Faker("en_IN")

        departments = ["CE", "ME", "Civil", "Electrical", "IT", "Chemical"]
        designations = [
            "Professor", "Associate Professor", "Assistant Professor",
            "Lab Assistant", "Head of Department", "Lecturer",
            "Workshop Instructor", "Librarian",
        ]
        years = ["FY", "SY", "TY"]

        added_students = 0
        added_staff = 0

        # ── Generate Students ────────────────────────────────
        for i in range(n_students):
            dept = random.choice(departments)
            year = random.choice(years)
            dept_code = dept[:2].upper() if len(dept) >= 2 else dept.upper()
            roll_no = f"24{dept_code}{i + 1:03d}"

            # Random issue date in last 18 months for realistic expiry spread
            issue_date = fake.date_between(start_date="-18m", end_date="today")

            ok, _ = add_student(
                name=fake.name(),
                roll_no=roll_no,
                department=dept,
                year=year,
                contact=fake.phone_number()[:10],
                email=fake.email(),
                photo_path="",
                issue_date=issue_date.strftime("%Y-%m-%d"),
            )
            if ok:
                added_students += 1

        # ── Generate Staff ───────────────────────────────────
        for i in range(n_staff):
            dept = random.choice(departments)
            designation = random.choice(designations)
            emp_id = f"EMP{i + 1:04d}"

            issue_date = fake.date_between(start_date="-18m", end_date="today")

            ok, _ = add_staff(
                name=fake.name(),
                employee_id=emp_id,
                department=dept,
                designation=designation,
                contact=fake.phone_number()[:10],
                email=fake.email(),
                photo_path="",
                issue_date=issue_date.strftime("%Y-%m-%d"),
            )
            if ok:
                added_staff += 1

        total = added_students + added_staff
        log_action("DEMO_DATA", "system", details=f"students={added_students}, staff={added_staff}")
        return True, (
            f"Demo data generated successfully!\n\n"
            f"  👩‍🎓 Students added: {added_students}\n"
            f"  👔 Staff added: {added_staff}\n"
            f"  📋 Total: {total}\n\n"
            f"Skipped records had duplicate roll/employee IDs."
        )
    except ImportError:
        return False, "Faker is not installed. Run: pip install faker"
    except Exception as e:
        return False, str(e)

